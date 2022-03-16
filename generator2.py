from openpyxl import load_workbook
from openpyxl.workbook import workbook
from subject import Subject
from teacher import teacher
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.workbook.workbook import Workbook
import matplotlib.pyplot as plt
import numpy as np

wb = load_workbook('nicein.xlsx')
wb.active = 0

sheet = wb.active

# First lets calculate how many students does the AUT exams have

columns = ['g', 'i']
aut_subjects = {}

for column in columns:
    i = 1
    while not sheet['a' + str(i)].value == None:
        subject = str(sheet[column + str(i)].value)

        if subject.startswith("BMEVIAU"):
            if( not subject in aut_subjects.keys()  ):
                aut_subjects[subject] = Subject()
                aut_subjects[subject].students = 0
                aut_subjects[subject].code = subject
                aut_subjects[subject].teachers = 0

            aut_subjects[subject].students = aut_subjects[subject].students + 1

        i = i + 1

# checkpoint

#print(aut_subjects)

# now lets create a list of teachers and populate them


teachers = {}

wb.active = 2

sheet = wb.active

columnIndex = 1
i = 1

curCell = sheet[get_column_letter(columnIndex) + str(i)]
#print("####################")


while not curCell.value == None:
    subject = curCell.value

    if not subject.startswith("BMEVIAU"):
        columnIndex = columnIndex + 1
        curCell = sheet[get_column_letter(columnIndex) + str(i)]
        continue

    #print("subject: " + subject)
    i = 3
    subject = curCell.value
    print("subject: " + subject + ", students: " + str(aut_subjects[subject].students))
    

    while not curCell.value == None: # ok this one looks scary AF, but trust me
        aut_subjects[subject].teachers = aut_subjects[subject].teachers + 1
        curCell = sheet[get_column_letter(columnIndex) + str(i)]

        teachername = curCell.value
        
        if(teachername == None):
            break
        #print("     " + teachername)
        #print(curCell.value)

        if( not teachername in teachers.keys()):
            teachers[teachername] = teacher(teachername)
            teachers[teachername].subjects = []
            teachers[teachername].relative_students = 0

        teachers[teachername].subjects.append(subject)
        print("     " + teachername + " " + subject)
        teachers[teachername].subjectcnt = teachers[teachername].subjectcnt + 1
        teachers[teachername].students = teachers[teachername].students + aut_subjects[subject].students

        i = i + 1 # next row
        curCell = sheet[get_column_letter(columnIndex) + str(i)]
    
#region thingies
    print("-------------")
    columnIndex = columnIndex + 1
    if(columnIndex > 5 ):
        print(teachers.keys())

        for teachername in teachers.keys():
            print(teachername)
            teachers[teachername].students = 0

            for subject in teachers[teachername].subjects:
                print("     " + subject)
                teachers[teachername].students = teachers[teachername].students + aut_subjects[subject].students
                
            print("--------------------")

        print(aut_subjects.keys())
        for key in aut_subjects.keys():
            print(key)
            print("With refer: " + aut_subjects[key].code + ": " + str(aut_subjects[key].teachers))

            subject = aut_subjects[key]
            print("With var: " + subject.code + ": " + str(subject.teachers))

        print(str(aut_subjects["BMEVIAUMA09"].teachers) + " " + str(aut_subjects["BMEVIAUMA09"].students))

        testteacher = teachers["Dr. Csorba Kristóf"]
        testteacher.calculateRelativeStudents(aut_subjects)
        print(testteacher.relative_students)
        #exit(1)
#endregion

    i = 1
    curCell = sheet[get_column_letter(columnIndex) + str(i)]
    #print(get_column_letter(columnIndex) + "|||")

#print("##############")
#teachers.keys()
#print("--##--##--##--##--")

for teachername in teachers.keys():
    #print(teachername)

    for subject in teachers[teachername].subjects:
        #print("     " + subject)
        #teachers[teachername].students = teachers[teachername].students + aut_subjects[subject]
        pass
    #print("--------------------")




#########################


for teachername in teachers.keys():
    teachers[teachername].calculateRelativeStudents(aut_subjects)


workbook2 = load_workbook("y.xlsx")
workbook2.active = 0
sheet = workbook2.active

# region test for a random teacher
testteacher = teachers["Dr. Dudás Ákos"] # type: teacher

testteacher.WorkHourWorkLoad([sheet],"j", "m","n",["o","p"],"q")

print (testteacher.workload)

print("--------")
testteacher.SummaWorkLoad([sheet],"j", "m","n",["o","p"],"q")
print("--------")
print(testteacher.workload)

print(testteacher.subjects)
print(testteacher.students)
print(aut_subjects[testteacher.subjects[0]].students)

#exit(1)
#endregion


for teachername in  teachers.keys():
    teachers[teachername].SummaWorkLoad([sheet],"j", "m","n",["o","p"],"q")
    print(teachername + ": " + str(teachers[teachername].workload) + " hrs, " +
    #str(teachers[teachername].subjectcnt) + " pcs subject, " + str(teachers[teachername].students) + " students")
    str(teachers[teachername].relative_students) + " rel student, " + str(teachers[teachername].students) + " students")

# CHECKPOINT


#region plotting
xs = []
ys = []
zs = []

for teacherName in teachers.keys():
    xs.append(teachers[teacherName].students)
    zs.append(teachers[teacherName].relative_students)
    ys.append(teachers[teacherName].workload)

x = np.array(xs)
y = np.array(ys)
z = np.array(zs)

fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')
ax.scatter(xs,zs,ys)
ax.set_ylim(0, 25)

plt.show()
#endregion