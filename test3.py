from interpolate import interpolate
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib import cm
from matplotlib.ticker import LinearLocator
import numpy as np
import math



wb = load_workbook('testinput.xlsx')
wb.active = 1
sheet = wb.active

print(wb.sheetnames)

#xletter = ["J", "K", "L", "M"]
#letters = ["B", "C", "D", "E", "F", "G", "H", "I"]

xletter = ["B", "C", "D", "E"]
letters = ["F", "G", "H", "I", "J", "K", "L", "M"]
numbers = [2,3]

xk = []
yk = []

for letter in letters:
    datapoint = []
    for number in numbers:
        cellid = letter + str(number)
        datapoint.append(sheet[cellid].value)
    xk.append(datapoint)
    print(datapoint)
    cellid = letter + "4"
    yk.append(sheet[cellid].value)



train_lectured = 0

for thingy in xk:
    train_lectured = train_lectured + thingy[0]

print(yk)
already_assigned = sum(yk)
student_number = math.ceil( train_lectured * 1.33 )
needtoassing = student_number - already_assigned


print("student number: " + str(student_number) )
print("already assigned: " + str(already_assigned))
print("need to assign: " + str(needtoassing))


print("-------------------")
print("######################")
print("-------------------")


interpolater = interpolate()
interpolater.setXK(xk)
interpolater.setYK(yk)
interpolater.setM(1)


x = []



for letter in xletter:
    datapoint = []
    for number in numbers:
        cellid = letter + str(number)
        datapoint.append(sheet[cellid].value)
    x.append(datapoint)

interpolated_values = []
for value in x:
    interpolated_values.append( interpolater.interpolate(value) )
print("M: 1")
print("Difference to goal: " + str( abs( needtoassing - sum(interpolated_values ) ) ) )
print(interpolated_values)
print("-------------------")


# testing different Ms
step = 0.5
M = 1
for i in range(1,5):
    interpolater.setM( M + i * step )
    interpolated_values = []
    for value in x:
        interpolated_values.append( interpolater.interpolate(value) )

    print("M: " + str(M + i * step) )
    print("Difference to goal: " + str( abs( needtoassing - sum(interpolated_values ) ) ) )
    print(interpolated_values)
    print("-------------------")
 


print("######################")
print("-------------------")


step = 0.1
for i in range(1,7):
    interpolater.setM( M + i * step )
    interpolated_values = []
    for value in x:
        interpolated_values.append( interpolater.interpolate(value) )

    print("M: " + str(M - i * step) )
    print("Difference to goal: " + str( abs( needtoassing - sum(interpolated_values ) ) ) )
    print(interpolated_values)
    print("-------------------")


